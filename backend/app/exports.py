"""Excel workbook and PDF report builders for nominations (admin-only exports).

Both are driven by the same aggregates as the analytics endpoint
(:func:`app.analytics.compute_nomination_analytics`) so the numbers always agree.

Security: every text cell written to the workbook is passed through
:func:`_xlsx_safe`, which neutralises spreadsheet formula injection the same way
the CSV export does — a value a spreadsheet would treat as a formula (``=``,
``+``, ``-``, ``@``, or a leading control char) is prefixed with an apostrophe.
"""
from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .analytics import compute_nomination_analytics
from .config import get_settings
from .models import Category, Nomination
from .schemas.form_schema import FormDefinition

EVENT_TITLE = "The Redemption City Awards of Excellence"

_GOLD = colors.HexColor("#B98D2F")
_GRAPHITE = colors.HexColor("#1F1B16")
_PAPER = colors.HexColor("#F5EFE3")
_RULE = colors.HexColor("#D9CEB8")

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_CONTROL_PREFIXES = ("\t", "\r", "\n")


# --- shared helpers -----------------------------------------------------------

def _xlsx_safe(value: Any) -> Any:
    """Return a spreadsheet-safe cell value (numbers pass through untouched)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(v) for v in value)
    text = str(value)
    if text[:1] in _FORMULA_PREFIXES or text[:1] in _CONTROL_PREFIXES:
        return "'" + text
    return text


def _esc(text: str) -> str:
    """Escape XML specials for reportlab Paragraph markup (Table cells are literal
    and don't need this, but Paragraph text does — e.g. 'Male & Female')."""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _sheet_title(name: str, used: set[str]) -> str:
    """A unique, Excel-legal sheet title (<=31 chars, none of []:*?/\\)."""
    cleaned = "".join(c for c in name if c not in set('[]:*?/\\')).strip() or "Category"
    cleaned = cleaned[:31]
    candidate, i = cleaned, 1
    while candidate.lower() in used:
        suffix = f" ({i})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate.lower())
    return candidate


# --- Excel --------------------------------------------------------------------

def build_nominations_xlsx(session: Session, category_slug: Optional[str] = None) -> bytes:
    analytics = compute_nomination_analytics(session, category_slug)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_summary_sheet(summary, analytics)

    cats_stmt = select(Category).order_by(Category.sort_order, Category.name)
    if category_slug:
        cats_stmt = cats_stmt.where(Category.slug == category_slug)
    categories = list(session.scalars(cats_stmt))

    used: set[str] = {"summary"}
    for cat in categories:
        noms = list(
            session.scalars(
                select(Nomination)
                .where(Nomination.category_id == cat.id)
                .order_by(Nomination.created_at.desc())
                .options(selectinload(Nomination.files))
            )
        )
        # Skip empty categories in the "all" export to avoid 20-odd blank tabs;
        # always include the sheet when a single category was requested.
        if not noms and not category_slug:
            continue
        _write_category_sheet(wb, cat, noms, used)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F1B16")
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autofit(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width, 10), 60)


def _write_summary_sheet(ws, a: dict[str, Any]) -> None:
    ws.append([EVENT_TITLE])
    ws["A1"].font = Font(bold=True, size=15, color="1F1B16")
    scope = "All categories" if a["scope"] == "all" else f"Category: {a['scope']}"
    ws.append([f"Nominations summary — {scope}"])
    ws.append([f"Generated {a['generated_at']}"])
    ws.append([])

    kpis = [
        ("Total nominations", a["total"]),
        ("Categories with entries", a["categories_with_entries"]),
        ("Empty categories", a["categories_empty"]),
        ("Unique nominators", a["unique_nominators"]),
        ("Nominations with evidence", a["with_evidence"]),
        ("Submitted in last 24h", a["last_24h"]),
        ("Submitted in last 7 days", a["last_7d"]),
        ("Submitted (pending review)", a["status"]["submitted"]),
        ("Shortlisted", a["status"]["shortlisted"]),
        ("Rejected", a["status"]["rejected"]),
    ]
    ws.append(["Metric", "Value"])
    metric_header_row = ws.max_row
    for label, value in kpis:
        ws.append([_xlsx_safe(label), _xlsx_safe(value)])
    for cell in ws[metric_header_row]:
        cell.font = Font(bold=True)

    ws.append([])
    ws.append(["Category", "Group", "Total", "Submitted", "Shortlisted", "Rejected"])
    table_header_row = ws.max_row
    for c in a["by_category"]:
        ws.append([
            _xlsx_safe(c["name"]), _xlsx_safe(c["group"]), c["count"],
            c["submitted"], c["shortlisted"], c["rejected"],
        ])
    for cell in ws[table_header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F1B16")
    _autofit(ws, {1: 40, 2: 16, 3: 10, 4: 12, 5: 12, 6: 10})


def _write_category_sheet(wb: Workbook, cat: Category, noms: list[Nomination], used: set[str]) -> None:
    ws = wb.create_sheet(_sheet_title(cat.name, used))
    try:
        fields = FormDefinition.model_validate(cat.form_schema).all_fields
    except Exception:
        fields = []

    meta_headers = ["ID", "Status", "Submitted (UTC)", "Nominator", "Contact", "Residency"]
    headers = meta_headers + [f.label for f in fields] + ["Evidence"]
    ws.append([_xlsx_safe(h) for h in headers])
    _style_header(ws, len(headers))
    ws.freeze_panes = "A2"

    for nom in noms:
        row: list[Any] = [
            nom.id,
            nom.status.value,
            nom.created_at.isoformat(sep=" ", timespec="minutes") if nom.created_at else "",
            nom.nominator_name or "",
            nom.nominator_contact or "",
            nom.residency or "",
        ]
        answers = nom.answers or {}
        row.extend(answers.get(f.key) for f in fields)
        row.append("; ".join(f.url for f in nom.files))
        ws.append([_xlsx_safe(v) for v in row])

    widths = {1: 8, 2: 12, 3: 18, 4: 22, 5: 24, 6: 14}
    for idx, f in enumerate(fields, start=len(meta_headers) + 1):
        widths[idx] = 28 if f.type.value in {"paragraph", "short_text"} else 16
    widths[len(headers)] = 40
    _autofit(ws, widths)


# --- PDF ----------------------------------------------------------------------

def build_nominations_pdf(session: Session, category_slug: Optional[str] = None) -> bytes:
    a = compute_nomination_analytics(session, category_slug)
    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        title="Nominations Report",
        author=EVENT_TITLE,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("rc_h1", parent=styles["Title"], textColor=_GRAPHITE, fontSize=22, spaceAfter=2)
    sub = ParagraphStyle("rc_sub", parent=styles["Normal"], textColor=_GOLD, fontSize=11, spaceAfter=1, leading=15)
    meta = ParagraphStyle("rc_meta", parent=styles["Normal"], textColor=colors.grey, fontSize=8.5, spaceAfter=10)
    h2 = ParagraphStyle("rc_h2", parent=styles["Heading2"], textColor=_GRAPHITE, fontSize=13, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("rc_body", parent=styles["Normal"], textColor=_GRAPHITE, fontSize=9.5, leading=13)

    edition = get_settings().edition_year
    scope = "All categories" if a["scope"] == "all" else f"Category — {a['scope']}"

    flow: list[Any] = [
        Paragraph(EVENT_TITLE, h1),
        Paragraph(f"Nominations Report · {edition} Edition", sub),
        Paragraph(_esc(scope), sub),
        Paragraph(f"Generated {a['generated_at']}", meta),
    ]

    # KPI grid (2 columns of label/value pairs)
    kpi_pairs = [
        ("Total nominations", a["total"]),
        ("Unique nominators", a["unique_nominators"]),
        ("Categories with entries", f"{a['categories_with_entries']} / {a['categories_total']}"),
        ("Empty categories", a["categories_empty"]),
        ("With evidence attached", a["with_evidence"]),
        ("Last 24 hours", a["last_24h"]),
        ("Submitted (pending)", a["status"]["submitted"]),
        ("Last 7 days", a["last_7d"]),
        ("Shortlisted", a["status"]["shortlisted"]),
        ("Rejected", a["status"]["rejected"]),
    ]
    kpi_rows = []
    for i in range(0, len(kpi_pairs), 2):
        left = kpi_pairs[i]
        right = kpi_pairs[i + 1] if i + 1 < len(kpi_pairs) else ("", "")
        kpi_rows.append([left[0], str(left[1]), right[0], str(right[1])])
    kpi_table = Table(kpi_rows, colWidths=[52 * mm, 22 * mm, 52 * mm, 22 * mm])
    kpi_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TEXTCOLOR", (0, 0), (-1, -1), _GRAPHITE),
        ("TEXTCOLOR", (1, 0), (1, -1), _GOLD),
        ("TEXTCOLOR", (3, 0), (3, -1), _GOLD),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [_PAPER, colors.white]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, _RULE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    flow.append(kpi_table)

    # By group
    if a["by_group"]:
        flow.append(Paragraph("By group", h2))
        group_rows = [["Group", "Nominations"]] + [
            [g["group"].title(), str(g["count"])] for g in a["by_group"]
        ]
        flow.append(_data_table(group_rows, [120 * mm, 40 * mm]))

    # Per-category breakdown
    flow.append(Paragraph("By category", h2))
    cat_rows = [["Category", "Total", "Submitted", "Shortlisted", "Rejected"]]
    for c in a["by_category"]:
        cat_rows.append([
            c["name"], str(c["count"]), str(c["submitted"]),
            str(c["shortlisted"]), str(c["rejected"]),
        ])
    flow.append(_data_table(cat_rows, [78 * mm, 20 * mm, 24 * mm, 24 * mm, 20 * mm]))

    if a["empty_categories"]:
        flow.append(Paragraph("Categories with no nominations yet", h2))
        flow.append(Paragraph(_esc(", ".join(a["empty_categories"])), body))

    doc.build(flow)
    return bio.getvalue()


def _data_table(rows: list[list[str]], col_widths: list[float]) -> Table:
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _GRAPHITE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 1), (-1, -1), _GRAPHITE),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _PAPER]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, _RULE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
    ]))
    return table
