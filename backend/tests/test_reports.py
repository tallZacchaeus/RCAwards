"""Nomination analytics + Excel/PDF export endpoints (admin only)."""
import io

import openpyxl

from tests.conftest import valid_creche_payload


def _seed(client, n=2):
    for _ in range(n):
        resp = client.post("/nominations", json=valid_creche_payload())
        assert resp.status_code == 201, resp.text


# --- analytics ----------------------------------------------------------------

def test_analytics_admin_only(client, judge_headers):
    assert client.get("/admin/analytics/nominations").status_code == 401
    assert client.get("/admin/analytics/nominations", headers=judge_headers).status_code == 403


def test_analytics_counts(client, admin_headers):
    _seed(client, 2)
    resp = client.get("/admin/analytics/nominations", headers=admin_headers)
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["scope"] == "all"
    assert data["total"] >= 2
    assert data["categories_total"] >= 1
    assert data["status"]["submitted"] >= 2
    creche = next((c for c in data["by_category"] if c["slug"] == "creche-of-the-year"), None)
    assert creche is not None and creche["count"] >= 2
    # sorted busiest-first
    counts = [c["count"] for c in data["by_category"]]
    assert counts == sorted(counts, reverse=True)


def test_analytics_category_scope_and_404(client, admin_headers):
    ok = client.get("/admin/analytics/nominations?category=creche-of-the-year", headers=admin_headers)
    assert ok.status_code == 200
    body = ok.json()
    assert body["scope"] == "creche-of-the-year"
    assert body["categories_total"] == 1
    missing = client.get("/admin/analytics/nominations?category=not-a-real-slug", headers=admin_headers)
    assert missing.status_code == 404


# --- Excel --------------------------------------------------------------------

def test_xlsx_export_shape(client, admin_headers):
    _seed(client, 2)
    resp = client.get("/admin/nominations/export.xlsx", headers=admin_headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.headers["content-disposition"].endswith("nominations-all.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Summary" in wb.sheetnames
    # at least one per-category sheet exists for the seeded nominations
    assert any(name != "Summary" for name in wb.sheetnames)


def test_xlsx_formula_injection_neutralised(client, admin_headers):
    payload = valid_creche_payload()
    payload["answers"]["creche_name"] = "=INJECT_TEST_123"
    assert client.post("/nominations", json=payload).status_code == 201
    resp = client.get(
        "/admin/nominations/export.xlsx?category=creche-of-the-year", headers=admin_headers
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    values = [c.value for ws in wb.worksheets for row in ws.iter_rows() for c in row]
    # the dangerous leading '=' is quoted, and the raw formula never appears
    assert "'=INJECT_TEST_123" in values
    assert "=INJECT_TEST_123" not in values


def test_xlsx_admin_only(client, judge_headers):
    assert client.get("/admin/nominations/export.xlsx").status_code == 401
    assert client.get("/admin/nominations/export.xlsx", headers=judge_headers).status_code == 403


# --- PDF ----------------------------------------------------------------------

def test_pdf_report(client, admin_headers):
    resp = client.get("/admin/reports/nominations.pdf", headers=admin_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"
    assert resp.headers["content-disposition"].endswith("nominations-report-all.pdf")


def test_pdf_admin_only(client, judge_headers):
    assert client.get("/admin/reports/nominations.pdf").status_code == 401
    assert client.get("/admin/reports/nominations.pdf", headers=judge_headers).status_code == 403
